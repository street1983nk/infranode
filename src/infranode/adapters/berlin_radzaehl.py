"""Berlin-Radzähldaten-Adapter ``fetch_berlin_radzaehl`` (DATA-40, Tier A).

Liefert den jüngsten Stundenwert je Berliner Radzählstelle (~30 Stationen)
keylos aus der offenen Gesamtdatei (DL-DE/Zero 2.0, SenMVKU, [VERIFIED 2026-06-23]):

  berlin.de/.../zaehlstellen-und-fahrradbarometer/gesamtdatei-stundenwerte.xlsx

Die XLSX (~18 MB) hat ein Sheet "Standortdaten" (Zählstelle-ID, Beschreibung,
Breiten-/Laengengrad) und je Jahr ein Sheet "Jahresdatei JJJJ" (Spalte 0 =
Stundenzeitstempel, Spalten 1..N = Zählwert je Station; Header trägt
"ID\nInbetriebnahme"). Es wird das Sheet mit dem GRÖSSTEN Jahr gelesen und daraus
die LETZTE Zeile mit Daten (= frischster Stundenwert) je Station genommen, gejoint
mit den Koordinaten aus "Standortdaten" über die Zählstelle-ID.

Performance: openpyxl ``read_only``/``data_only`` streamt; der 18-MB-Fetch+Parse
läuft nur bei Cache-Miss (sehr lange TTL, ``_SOURCE_TTL["berlin_radzaehl"]``).

Sicherheit (T-9-02 SSRF): Host hartkodiert. DoS-/Datenfehler-Schutz:
``raise_for_status()`` (5xx -> STALE-ON-ERROR der Fassade); Felder defensiv.
"""

from __future__ import annotations

import io
from datetime import datetime

import httpx
import openpyxl

_XLSX_URL = (
    "https://www.berlin.de/sen/uvk/_assets/verkehr/verkehrsplanung/radverkehr/"
    "weitere-radinfrastruktur/zaehlstellen-und-fahrradbarometer/"
    "gesamtdatei-stundenwerte.xlsx"
)
_STANDORT_SHEET = "Standortdaten"
_YEAR_SHEET_PREFIX = "Jahresdatei "


def _standorte(wb) -> dict[str, dict]:
    """``Standortdaten``-Sheet -> {zaehlstelle_id: {name, lat, lon}}."""
    if _STANDORT_SHEET not in wb.sheetnames:
        return {}
    out: dict[str, dict] = {}
    rows = wb[_STANDORT_SHEET].iter_rows(values_only=True)
    next(rows, None)  # Kopfzeile
    for row in rows:
        if not row or not row[0]:
            continue
        try:
            lat = float(row[2]) if len(row) > 2 and row[2] is not None else None
            lon = float(row[3]) if len(row) > 3 and row[3] is not None else None
        except (TypeError, ValueError):
            lat = lon = None
        out[str(row[0]).strip()] = {
            "name": row[1] if len(row) > 1 else None,
            "lat": lat,
            "lon": lon,
        }
    return out


def _latest_year_sheet(wb) -> str | None:
    """Name des ``Jahresdatei``-Sheets mit dem größten Jahr (oder None)."""
    best: tuple[int, str] | None = None
    for name in wb.sheetnames:
        if not name.startswith(_YEAR_SHEET_PREFIX):
            continue
        try:
            year = int(name[len(_YEAR_SHEET_PREFIX) :].strip())
        except ValueError:
            continue
        if best is None or year > best[0]:
            best = (year, name)
    return best[1] if best else None


def _latest_counts(wb, sheet_name: str) -> tuple[dict[str, int], str | None]:
    """Letzte Datenzeile des Jahres-Sheets -> ({zaehlstelle_id: value}, period_iso)."""
    rows = wb[sheet_name].iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {}, None
    # Spaltenindex (>=1) -> Zählstelle-ID (Header "ID\nInbetriebnahme").
    col_id: dict[int, str] = {}
    for idx, cell in enumerate(header):
        if idx == 0 or not cell:
            continue
        col_id[idx] = str(cell).split("\n")[0].strip()
    last_values: dict[str, int] = {}
    last_period: str | None = None
    for row in rows:
        if not row or not isinstance(row[0], datetime):
            continue
        values: dict[str, int] = {}
        for idx, sid in col_id.items():
            if idx < len(row) and isinstance(row[idx], int | float):
                values[sid] = int(row[idx])
        if values:
            last_values = values
            last_period = row[0].isoformat()
    return last_values, last_period


async def fetch_berlin_radzaehl(
    http: httpx.AsyncClient,
    *,
    slug: str,
    lat: float,
    lon: float,
    radius_km: float = 30.0,
) -> dict:
    """Holt die jüngsten Berliner Rad-Stundenwerte (XLSX-Gesamtdatei).

    GET der XLSX, openpyxl read_only: Standorte (Koordinaten) + neuestes
    Jahres-Sheet (letzte Datenzeile je Station). Join über die Zählstelle-ID.
    ``lat``/``lon``/``radius_km`` sind vertragskonform Teil der Signatur (ungenutzt;
    Berlin liefert den kompletten Stadt-Datensatz).

    Rückgabe-Keys (exakt das, was ``map_berlin_radzaehl`` erwartet): ``slug``,
    ``stations`` (je Station name/id/lat/lon/value/period) und ``as_of``.
    """
    resp = await http.get(_XLSX_URL)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(
        io.BytesIO(resp.content), read_only=True, data_only=True
    )
    try:
        standorte = _standorte(wb)
        sheet = _latest_year_sheet(wb)
        if sheet is None:
            return {"slug": slug, "stations": [], "as_of": None}
        values, period = _latest_counts(wb, sheet)
    finally:
        wb.close()

    if not values:
        return {"slug": slug, "stations": [], "as_of": None}

    stations: list[dict] = []
    for sid, value in values.items():
        meta = standorte.get(sid, {})
        stations.append(
            {
                "station": meta.get("name") or sid,
                "station_id": sid,
                "lat": meta.get("lat"),
                "lon": meta.get("lon"),
                "value": value,
                "period": period,
            }
        )
    return {"slug": slug, "stations": stations, "as_of": period}
